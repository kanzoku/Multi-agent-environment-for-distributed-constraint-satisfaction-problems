import pandas as pd
from multiprocessing import Lock
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook


class TestResultsManager:
    def __init__(self, file_path):
        self.file_path = file_path
        self.sheet_name_overview = "Übersicht"
        self.sheet_name_detailed = "Ergebnisse"
        self.lock = Lock()
        self.df_test_series, self.df_detailed_test_series = self.load_dataframes()

    def load_dataframes(self):
        df_test_series = pd.read_excel(self.file_path, sheet_name=self.sheet_name_overview)
        df_detailed_test_series = pd.read_excel(self.file_path, sheet_name=self.sheet_name_detailed)
        return df_test_series, df_detailed_test_series

    def save_dataframes(self):
        with self.lock:
            # Laden Sie das bestehende Workbook
            workbook = load_workbook(self.file_path)

            # Funktion zum Überschreiben eines Sheets mit einem DataFrame
            def overwrite_sheet(sheet: Worksheet, dataframe: pd.DataFrame):
                dataframe = dataframe.astype(str).replace('None', '').replace({None: ''})
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.value = None  # Clear all existing cells
                for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        sheet.cell(row=r_idx, column=c_idx, value=value)

            # Überschreiben Sie die Sheets oder erstellen Sie sie, falls sie nicht existieren
            if self.sheet_name_overview in workbook.sheetnames:
                sheet_overview = workbook[self.sheet_name_overview]
            else:
                sheet_overview = workbook.create_sheet(title=self.sheet_name_overview)
            overwrite_sheet(sheet_overview, self.df_test_series)

            if self.sheet_name_detailed in workbook.sheetnames:
                sheet_detailed = workbook[self.sheet_name_detailed]
            else:
                sheet_detailed = workbook.create_sheet(title=self.sheet_name_detailed)
            overwrite_sheet(sheet_detailed, self.df_detailed_test_series)

            # Speichern Sie das Workbook
            workbook.save(self.file_path)
        print("Daten erfolgreich aktualisiert und gespeichert.")

    def add_test_series(self, data):
        with self.lock:
            new_row = pd.DataFrame([data])
            self.df_test_series = pd.concat([self.df_test_series, new_row], ignore_index=True)

    def add_detailed_test_series(self, data):
        with self.lock:
            new_row = pd.DataFrame([data])
            self.df_detailed_test_series = pd.concat([self.df_detailed_test_series, new_row], ignore_index=True)

    def update_test_series(self, test_id, updates):
        with self.lock:
            self.df_test_series.loc[self.df_test_series['Testreihe-ID'] == test_id, list(updates.keys())] = list(
                updates.values())

    def update_detailed_test_series(self, test_id, solution_id, updates):
        with self.lock:
            self.df_detailed_test_series.loc[
                (self.df_detailed_test_series['Testreihe-ID'] == test_id) & (
                            self.df_detailed_test_series['Lösung-ID'] == solution_id),
                list(updates.keys())
            ] = list(updates.values())

    def get_current_test_series(self):
        with self.lock:
            if self.df_test_series.empty:
                return None
            self.df_test_series['Testreihe-ID'] = self.df_test_series['Testreihe-ID'].astype(int)
            max_test_series_id = self.df_test_series['Testreihe-ID'].max()
        return self.df_test_series.loc[self.df_test_series['Testreihe-ID'] == max_test_series_id]

    def get_next_test_series_id(self):
        with self.lock:
            if self.df_test_series.empty:
                return 1
            return self.df_test_series['Testreihe-ID'].max() + 1

    def read_sudoku_to_dict(self, file_name, sheet_name, row_number):
        df = pd.read_excel(file_name, sheet_name=sheet_name)
        sudoku_row = df.iloc[row_number - 1].tolist()
        if sheet_name == "4x4":
            column_labels = [f"{chr(97 + i)}{j + 1}" for i in range(4) for j in range(4)]
        else:
            column_labels = [f"{chr(97 + i)}{j + 1}" for i in range(9) for j in range(9)]
        occupation_dict = {label: None for label in column_labels}
        for label, value in zip(column_labels, sudoku_row):
            if value == 0:
                occupation_dict[label] = None
            else:
                occupation_dict[label] = value
        return occupation_dict

    def read_sudoku(self, number, size, level):
        row_number = number + ((level - 1) * 100)
        occupation_dict = self.read_sudoku_to_dict(self.file_path, size, row_number)
        print(f"Das Sudoku in Zeile {row_number + 1} wurde erfolgreich eingelesen.")
        print(occupation_dict)
        return occupation_dict
